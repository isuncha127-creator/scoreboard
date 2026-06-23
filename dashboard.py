import re
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import warnings
import os
import requests
import urllib3
from datetime import datetime, timedelta, date
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings("ignore")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

st.set_page_config(
    page_title="03Y51 스코어보드",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

FILE_PATH = os.path.join(os.path.dirname(__file__), "03Y51 운용파일_20260602(6월)_mk.xlsx")

FACTOR_COLORS = {
    "B": "#4C72B0",
    "C": "#DD8452",
    "Q": "#55A868",
    "M": "#C44E52",
    "최종_S": "#8172B2",
}

SECTOR_ORDER = [
    "Information Technology",
    "Communication Services",
    "Consumer Discretionary",
    "Consumer Staples",
    "Financials",
    "Health Care",
    "Industrials",
    "Materials",
    "Real Estate",
    "Energy",
    "Utilities",
]

# Bloomberg country suffix → Yahoo Finance suffix
SUFFIX_MAP = {
    "US": "", "FP": ".PA", "GR": ".DE", "GY": ".DE", "SW": ".SW",
    "LN": ".L", "JP": ".T", "HK": ".HK", "KS": ".KS", "NA": ".AS",
    "IM": ".MI", "SM": ".MC", "AV": ".VI", "SS": ".SS", "AU": ".AX",
    "CN": ".TO", "SP": ".SI", "TW": ".TW", "NO": ".OL", "DC": ".CO",
    "SJ": ".JO", "MK": ".KL", "TB": ".BK", "IB": ".JK", "IN": ".NS",
}

# Hardcoded Yahoo tickers for stocks not mappable via Bloomberg suffix
ISIN_YAHOO_FALLBACK = {
    "KR7000660001": "000660.KS",  # SK Hynix
    "GB0009895292": "AZN",        # AstraZeneca (NASDAQ)
    "GB000989529B": "AZN",        # AstraZeneca (파일 내 ISIN 변형)
    "US0420682058": "ARM",        # ARM Holdings
    "US24703L2025": "DELL",       # Dell Technologies
    "US5128073062": "LRCX",       # Lam Research
    "FR0000121329": "HO.PA",      # Thales SA
    "US92537N1081": "VRT",        # Vertiv Holdings
    "US1717793095": "CIEN",       # Ciena
    "BRABEVACNOR1": "ABEV3.SA",   # Ambev (B3)
    "US88033G4073": "THC",        # Tenet Healthcare
    "US2787681061": "SATS",       # EchoStar
    "US7731211089": "RKLB",       # Rocket Lab
    "US00217D1000": "ASTS",       # AST SpaceMobile
    "US72703X1063": "PL",         # Planet Labs
}

YAHOO_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


def bbg_to_yahoo(bbg: str) -> str:
    parts = bbg.strip().split()
    if len(parts) < 2:
        return re.sub(r"[^A-Za-z0-9\-\.]", "", parts[0])
    ticker = re.sub(r"[^A-Za-z0-9]", "", parts[0])  # "/" 등 특수문자 제거 (e.g. "RR/")
    country = parts[1].upper()
    return ticker + SUFFIX_MAP.get(country, f".{country}")


# ─── 데이터 로딩 ──────────────────────────────────────────────────────────────

@st.cache_data(ttl=60, show_spinner="엑셀 파일 로딩 중...")
def load_data():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)

    # ── 1. KPI (최종스코어 상단 요약) ──
    ws_score = wb["최종스코어"]
    all_rows = list(ws_score.iter_rows(values_only=True))

    kpi = {}
    kpi_map = {
        0: ("리밸_종목수", None),
        1: ("현재_종목수", None),
        2: ("Inter_복제율", 0.50),
        3: ("ExPort_복제율", None),
        4: ("턴오버", None),
        5: ("인터브랜드_비중", 0.60),
        6: ("BM_복제율", 0.30),
    }
    for row_i, (key, target) in kpi_map.items():
        val = all_rows[row_i][1]
        kpi[key] = {"value": val, "target": target}

    # ── 2. 종목 데이터 (행 12~, 헤더 행 11) ──
    # 컬럼 인덱스 매핑
    COL = {
        "memo": 0, "ex_diff": 1, "ISIN": 4, "Name": 5,
        "GICS": 6, "IndustryGroup": 7, "Industry": 8, "Country": 9,
        "MCW_Inter": 10, "EW_Inter": 11, "BM_W": 12,
        "EX_W": 13, "EX_AW": 14,
        "Total_B": 15, "B": 16,
        "Total_C": 17, "C": 19,
        "Total_Q": 20, "Q": 22,
        "Total_M": 23, "M": 25,
        "Final_S": 26,
        "선택": 31,
        "1단계포트": 35,
        "최종포트": 38,
        "최종AW": 39,
        "시가총액": 47,
    }

    records = []
    for row in all_rows[11:]:
        isin = row[COL["ISIN"]]
        if not isinstance(isin, str):
            continue
        isin = isin.strip()
        if not isin or len(isin) < 12:
            continue

        def safe(col):
            v = row[col]
            return v if v not in ("#VALUE!", "#NAME?", "#REF!", "#DIV/0!") else None

        def safe_str(col):
            v = safe(col)
            return v if isinstance(v, str) else None

        records.append({
            "ISIN": isin,
            "Name": safe(COL["Name"]),
            "GICS": safe_str(COL["GICS"]),
            "IndustryGroup": safe_str(COL["IndustryGroup"]),
            "Industry": safe_str(COL["Industry"]),
            "Country": safe_str(COL["Country"]),
            "MCW_Inter": safe(COL["MCW_Inter"]),
            "EW_Inter": safe(COL["EW_Inter"]),
            "BM_W": safe(COL["BM_W"]),
            "EX_W": safe(COL["EX_W"]),
            "EX_AW": safe(COL["EX_AW"]),
            "Total_B": safe(COL["Total_B"]),
            "B": safe(COL["B"]),
            "Total_C": safe(COL["Total_C"]),
            "C": safe(COL["C"]),
            "Total_Q": safe(COL["Total_Q"]),
            "Q": safe(COL["Q"]),
            "Total_M": safe(COL["Total_M"]),
            "M": safe(COL["M"]),
            "Final_S": safe(COL["Final_S"]),
            "선택": safe(COL["선택"]),
            "최종포트": safe(COL["최종포트"]),
            "최종AW": safe(COL["최종AW"]),
            "시가총액": safe(COL["시가총액"]),
        })

    df = pd.DataFrame(records)
    df = df[df["Name"].notna()].copy()

    # 숫자형 변환
    num_cols = ["MCW_Inter", "EW_Inter", "BM_W", "EX_W", "EX_AW",
                "Total_B", "B", "Total_C", "C", "Total_Q", "Q", "Total_M", "M",
                "Final_S", "선택", "최종포트", "최종AW", "시가총액"]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # 최종포트 > 0 이면 편입 (KPI 리밸 종목수 62와 일치)
    df["편입"] = df["최종포트"].fillna(0) > 0

    # ── 3. AGG&ECALL 테마 ──
    ws_agg = wb["AGG&ECALL"]
    agg_rows = list(ws_agg.iter_rows(values_only=True))

    # 테마 컬럼 시작 위치 (간격 7~8)
    theme_starts = []
    for i, v in enumerate(agg_rows[0]):
        if isinstance(v, str) and v.strip():
            theme_starts.append((i, v.strip()))

    themes = {}
    for t_idx, (col_start, theme_name) in enumerate(theme_starts):
        # 다음 테마 시작 컬럼
        next_start = theme_starts[t_idx + 1][0] if t_idx + 1 < len(theme_starts) else len(agg_rows[0])

        # 헤더 행(행2): STOCK, AGG, ECALL, ISIN 컬럼 위치 파악
        header_row = agg_rows[1]
        col_stock = col_start + 1 if col_start + 1 < next_start else None
        col_agg = col_start + 2 if col_start + 2 < next_start else None
        col_ecall = col_start + 3 if col_start + 3 < next_start else None
        col_isin = col_start + 4 if col_start + 4 < next_start else None

        stocks = []
        for row in agg_rows[2:]:
            stock_name = row[col_stock] if col_stock is not None else None
            if not isinstance(stock_name, str) or not stock_name.strip():
                continue
            agg_val = row[col_agg] if col_agg is not None else None
            ecall_val = row[col_ecall] if col_ecall is not None else None
            isin_val = row[col_isin] if col_isin is not None else None
            stocks.append({
                "Name": stock_name.strip(),
                "AGG": agg_val,
                "ECALL": ecall_val,
                "ISIN": isin_val,
            })

        # AGG 집계값
        agg_count = agg_rows[0][col_start + 2] if col_start + 2 < len(agg_rows[0]) else None
        agg_ecall = agg_rows[0][col_start + 3] if col_start + 3 < len(agg_rows[0]) else None

        themes[theme_name] = {
            "stocks": pd.DataFrame(stocks),
            "agg_count": agg_count,
            "agg_ecall": agg_ecall,
        }

    # ── 4. 팩터 세부 데이터 ──
    factor_detail = {}

    # 브랜드 특화 스코어
    ws_b = wb["1. 브랜드 특화 스코어"]
    rows_b = list(ws_b.iter_rows(values_only=True))
    b_records = []
    for row in rows_b[11:]:
        isin = row[0]
        if not isinstance(isin, str) or len(isin.strip()) < 12:
            continue
        b_records.append({
            "ISIN": isin.strip(),
            "인터브랜드포함": row[2],
            "브랜드가치성장(BV_S)": row[10],
            "중국CNPP포함": row[11],
            "중국CNPP스코어": row[12],
            "중국SCORE": row[13],
            "웰링턴/ETF포함": row[14],
            "Total_B": row[15],
        })
    df_b = pd.DataFrame(b_records)
    for c in ["인터브랜드포함","브랜드가치성장(BV_S)","중국CNPP포함","중국CNPP스코어","중국SCORE","웰링턴/ETF포함","Total_B"]:
        df_b[c] = pd.to_numeric(df_b[c], errors="coerce")
    factor_detail["B"] = {
        "df": df_b,
        "items": ["브랜드가치성장(BV_S)", "중국CNPP스코어", "웰링턴/ETF포함"],
        "weights": {"브랜드가치성장(BV_S)": 1.0, "중국CNPP스코어": 0.0, "웰링턴/ETF포함": 0.0},
        "label": "1. 브랜드 특화 스코어",
    }

    # 경쟁력 스코어
    ws_c = wb["2. 경쟁력 스코어"]
    rows_c = list(ws_c.iter_rows(values_only=True))
    c_records = []
    for row in rows_c[11:]:
        isin = row[0]
        if not isinstance(isin, str) or len(isin.strip()) < 12:
            continue
        c_records.append({
            "ISIN": isin.strip(),
            "WIDE MOAT": row[2],
            "무형자산": row[3],
            "네트워크효과": row[4],
            "규모의경제": row[5],
            "비용우위": row[6],
            "전환비용": row[7],
            "Moat_S": row[8],
            "AI/반도체": row[9],
            "우주/방산": row[10],
            "은행": row[11],
            "웰링턴퀄리티": row[12],
            "럭셔리": row[13],
            "Total_C": row[14],
        })
    df_c = pd.DataFrame(c_records)
    for c in ["WIDE MOAT","무형자산","네트워크효과","규모의경제","비용우위","전환비용","Moat_S","AI/반도체","우주/방산","은행","웰링턴퀄리티","럭셔리","Total_C"]:
        df_c[c] = pd.to_numeric(df_c[c], errors="coerce")
    factor_detail["C"] = {
        "df": df_c,
        "items": ["Moat_S", "AI/반도체", "우주/방산", "은행", "럭셔리"],
        "weights": {"Moat_S": 0.6, "AI/반도체": 0.1, "우주/방산": 0.1, "은행": 0.1, "럭셔리": 0.1},
        "moat_items": ["무형자산", "네트워크효과", "규모의경제", "비용우위", "전환비용"],
        "label": "2. 경쟁력 스코어",
    }

    # Traditional Q
    ws_q = wb["3. Traditional Q"]
    rows_q = list(ws_q.iter_rows(values_only=True))
    q_records = []
    for row in rows_q[11:]:
        isin = row[0]
        if not isinstance(isin, str) or len(isin.strip()) < 12:
            continue
        q_records.append({
            "ISIN": isin.strip(),
            "시가총액": row[2],
            "현재가": row[4],
            "가격_1Y": row[5],
            "가격_YTD": row[6],
            "가격_3M": row[7],
            "가격_1M": row[8],
            "가격_1W": row[9],
            "EQ(Earning Quality)": row[10],
            "FCF마진3Y": row[11],
            "Q(Quality)": row[12],
            "Morningstar Rating": row[13],
            "DDM괴리율": row[15],
            "업종대비RelVal": row[16],
            "P/BV자기대비": row[17],
            "V(Valuation)": row[18],
            "Sentiment(ARM)": row[19],
            "Total_Q": row[12],
        })
    df_q = pd.DataFrame(q_records)
    for c in ["시가총액","현재가","가격_1Y","가격_YTD","가격_3M","가격_1M","가격_1W",
              "EQ(Earning Quality)","FCF마진3Y","Q(Quality)","Morningstar Rating",
              "DDM괴리율","업종대비RelVal","P/BV자기대비","V(Valuation)","Sentiment(ARM)"]:
        df_q[c] = pd.to_numeric(df_q[c], errors="coerce")

    # 수익률 계산 (가격 기반)
    def calc_ret(cur, ref):
        return (cur - ref) / ref if (pd.notna(ref) and ref > 0 and pd.notna(cur)) else None

    df_q["YTD_R"] = df_q.apply(lambda r: calc_ret(r["현재가"], r["가격_YTD"]), axis=1)
    df_q["1M_R"]  = df_q.apply(lambda r: calc_ret(r["현재가"], r["가격_1M"]),  axis=1)
    df_q["1W_R"]  = df_q.apply(lambda r: calc_ret(r["현재가"], r["가격_1W"]),  axis=1)
    factor_detail["Q"] = {
        "df": df_q,
        "items": ["EQ(Earning Quality)", "Q(Quality)", "Morningstar Rating", "V(Valuation)", "Sentiment(ARM)"],
        "weights": {"EQ(Earning Quality)": 0.2, "Q(Quality)": 0.2, "Morningstar Rating": 0.2, "V(Valuation)": 0.2, "Sentiment(ARM)": 0.2},
        "label": "3. Traditional Quality",
    }

    # Macro
    ws_m = wb["4. Macro"]
    rows_m = list(ws_m.iter_rows(values_only=True))
    m_records = []
    for row in rows_m[11:]:
        isin = row[0]
        if not isinstance(isin, str) or len(isin.strip()) < 12:
            continue
        m_records.append({
            "ISIN": isin.strip(),
            "AI빅테크/반도체": row[4],
            "AI활용": row[5],
            "우주/방산": row[6],
            "럭셔리": row[7],
            "금융": row[8],
            "중국빅테크": row[9],
            "소프트웨어": row[10],
            "Edge_S": row[11],
            "QLI": row[12],
            "GQR": row[13],
            "EMP_S": row[14],
            "Event_S": row[16],
            "Total_M": row[19],
        })
    df_m = pd.DataFrame(m_records)
    for c in ["AI빅테크/반도체","AI활용","우주/방산","럭셔리","금융","중국빅테크","소프트웨어","Edge_S","QLI","GQR","EMP_S","Event_S","Total_M"]:
        df_m[c] = pd.to_numeric(df_m[c], errors="coerce")
    factor_detail["M"] = {
        "df": df_m,
        "items": ["Edge_S", "EMP_S", "Event_S"],
        "weights": {"Edge_S": 0.4, "EMP_S": 0.3, "Event_S": 0.3},
        "theme_items": ["AI빅테크/반도체", "AI활용", "우주/방산", "럭셔리", "금융", "중국빅테크", "소프트웨어"],
        "label": "4. Macro",
    }

    # ── 5. PORT 시트에서 Bloomberg 티커 매핑 ──
    # PORT 시트 구조: col1=Bloomberg ticker, col2=ISIN, 데이터 row8~
    isin_yahoo_map = dict(ISIN_YAHOO_FALLBACK)  # fallback으로 초기화

    port_ws = None
    for sname in wb.sheetnames:
        su = sname.strip().upper()
        if su == "PORT" or su.startswith("PORT"):
            port_ws = wb[sname]
            break

    if port_ws is not None:
        port_rows = list(port_ws.iter_rows(values_only=True))
        for row in port_rows[8:]:  # 데이터는 row8(인덱스)부터
            if not row or len(row) < 3:
                continue
            bbg = row[1]
            isin = row[2]
            if not isinstance(bbg, str) or not isinstance(isin, str):
                continue
            isin = isin.strip()
            bbg = bbg.strip()
            if len(isin) != 12 or not bbg:
                continue
            if isin not in isin_yahoo_map:  # fallback 우선
                isin_yahoo_map[isin] = bbg_to_yahoo(bbg)

    factor_detail["ticker_map"] = isin_yahoo_map

    # ── 5. GroupBy_2: 섹터/테마/산업군/국가 비중 분석 ──
    ws_gb2 = wb["GroupBy_2"]
    gb2_rows = list(ws_gb2.iter_rows(values_only=True))

    def gb_block(start, name_col, count_labels, count_cols, pct_labels, pct_cols):
        recs = []
        for row in gb2_rows[start:]:
            name = row[name_col]
            if not isinstance(name, str) or not name.strip():
                break
            rec = {"항목": name.strip()}
            for label, c in zip(count_labels, count_cols):
                rec[label] = row[c]
            for label, c in zip(pct_labels, pct_cols):
                rec[label] = row[c]
            recs.append(rec)
        return pd.DataFrame(recs)

    groupby2 = {
        "sector": gb_block(
            4, 3, ["XPORT", "PORT", "차이"], [0, 1, 2],
            ["PORT비중", "XPORT비중", "MCW", "EW", "URTH", "Active", "XActive"], [4, 5, 6, 7, 8, 9, 10],
        ),
        "theme": gb_block(
            20, 3, ["XPORT", "PORT", "차이"], [0, 1, 2],
            ["PORT비중", "XPORT비중", "MCW", "EW", "URTH", "Active", "XActive"], [4, 5, 6, 7, 8, 9, 10],
        ),
        "industry_group": gb_block(
            4, 15, ["XPORT", "PORT", "차이"], [12, 13, 14],
            ["PORT비중", "XPORT비중", "JCW", "EW", "URTH", "Active", "XActive"], [16, 17, 18, 19, 20, 21, 22],
        ),
        "country": gb_block(
            4, 33, [], [],
            ["PORT비중", "XPORT비중", "URTH", "Active", "XActive"], [34, 35, 36, 37, 38],
        ),
    }

    wb.close()
    return df, kpi, themes, factor_detail, groupby2


# ─── 라이브 수익률 ────────────────────────────────────────────────────────────

def _fetch_one_ticker(isin: str, ticker: str, session: requests.Session) -> dict:
    try:
        url = (
            f"https://query2.finance.yahoo.com/v8/finance/chart/{ticker}"
            f"?interval=1d&range=1y"
        )
        r = session.get(url, timeout=12)
        if r.status_code != 200:
            return {"isin": isin, "error": f"HTTP {r.status_code}"}

        data = r.json()
        chart = data.get("chart", {}).get("result", [])
        if not chart:
            return {"isin": isin, "error": "no data"}

        timestamps = chart[0].get("timestamp", [])
        closes_raw = chart[0]["indicators"]["quote"][0].get("close", [])

        # None 제거
        pairs = [(t, c) for t, c in zip(timestamps, closes_raw) if c is not None]
        if len(pairs) < 2:
            return {"isin": isin, "error": "insufficient data"}

        last_ts, last_close = pairs[-1]
        prev_ts, prev_close = pairs[-2]

        last_dt = datetime.utcfromtimestamp(last_ts)

        def find_ref(days_ago):
            target = last_dt - timedelta(days=days_ago)
            best = None
            for ts, cl in pairs[:-1]:
                if datetime.utcfromtimestamp(ts) <= target:
                    best = cl
            return best

        # YTD 기준: 전년도 마지막 거래일 종가
        year_start = datetime(last_dt.year, 1, 1)
        ytd_ref_close = None
        for ts, cl in reversed(pairs[:-1]):
            if datetime.utcfromtimestamp(ts) < year_start:
                ytd_ref_close = cl
                break

        def pct(ref_close):
            if ref_close is None or ref_close <= 0:
                return None
            return (last_close - ref_close) / ref_close

        return {
            "isin": isin,
            "ticker": ticker,
            "last_price": last_close,
            "D":   pct(prev_close),
            "1W":  pct(find_ref(7)),
            "1M":  pct(find_ref(31)),
            "YTD": pct(ytd_ref_close),
            "error": None,
        }
    except Exception as e:
        return {"isin": isin, "error": str(e)[:80]}


@st.cache_data(ttl=900, show_spinner=False)
def fetch_live_returns(ticker_tuples: tuple) -> dict:
    """ticker_tuples: tuple of (isin, yahoo_ticker). Returns {isin: {...}}"""
    session = requests.Session()
    session.headers.update(YAHOO_HEADERS)
    session.verify = False

    result = {}
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {
            pool.submit(_fetch_one_ticker, isin, ticker, session): isin
            for isin, ticker in ticker_tuples
        }
        for fut in as_completed(futures):
            rec = fut.result()
            result[rec["isin"]] = rec

    return result


@st.cache_data(ttl=900, show_spinner=False)
def fetch_stock_news(ticker: str, count: int = 3) -> list:
    try:
        session = requests.Session()
        session.headers.update(YAHOO_HEADERS)
        session.verify = False
        r = session.get(
            f"https://query1.finance.yahoo.com/v1/finance/search?q={ticker}&newsCount={count}&quotesCount=0",
            timeout=10,
        )
        if r.status_code != 200:
            return []
        data = r.json()
        return [
            {"title": n.get("title"), "link": n.get("link"), "publisher": n.get("publisher")}
            for n in data.get("news", [])[:count]
        ]
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_company_profile(ticker: str) -> dict:
    try:
        session = requests.Session()
        session.headers.update(YAHOO_HEADERS)
        session.verify = False
        session.get("https://fc.yahoo.com", timeout=8)
        crumb = session.get("https://query1.finance.yahoo.com/v1/test/getcrumb", timeout=8).text.strip()
        r = session.get(
            f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{ticker}"
            f"?modules=assetProfile&crumb={crumb}",
            timeout=10,
        )
        if r.status_code != 200:
            return {}
        result = r.json().get("quoteSummary", {}).get("result") or []
        if not result:
            return {}
        prof = result[0].get("assetProfile", {})
        return {"summary": prof.get("longBusinessSummary"), "industry": prof.get("industry")}
    except Exception:
        return {}


def top_moat_factor(factor_detail, isin):
    df_c = factor_detail["C"]["df"]
    row = df_c[df_c["ISIN"] == isin]
    if row.empty:
        return None
    row = row.iloc[0]
    moat_items = ["무형자산", "네트워크효과", "규모의경제", "비용우위", "전환비용"]
    vals = {k: row[k] for k in moat_items if pd.notna(row.get(k)) and row[k] > 0}
    return max(vals, key=vals.get) if vals else None


# ─── 포맷 헬퍼 ───────────────────────────────────────────────────────────────

def fmt_pct(v, decimals=1):
    if v is None or pd.isna(v):
        return "N/A"
    return f"{v * 100:.{decimals}f}%"


def score_bar_color(v):
    if pd.isna(v):
        return ""
    if v >= 0.8:
        return "background-color:#c6efce;color:#276221"
    if v >= 0.6:
        return "background-color:#ffeb9c;color:#9c6500"
    return "background-color:#ffc7ce;color:#9c0006"


# ─── 탭 렌더링 함수들 ─────────────────────────────────────────────────────────

def render_gb_table(label, gb_df, count_cols, bar_cols):
    st.markdown(f"**{label}**")
    if gb_df.empty:
        st.caption("데이터 없음")
        return
    disp = gb_df.copy()
    pct_cols = [c for c in disp.columns if c != "항목" and c not in count_cols]
    for c in pct_cols:
        disp[c] = disp[c] * 100

    col_cfg = {}
    for c in count_cols + pct_cols:
        fmt = "%d" if c in count_cols else "%.1f%%"
        if c in bar_cols:
            vmax = float(disp[c].abs().max() or 1)
            col_cfg[c] = st.column_config.ProgressColumn(format=fmt, min_value=-vmax, max_value=vmax)
        else:
            col_cfg[c] = st.column_config.NumberColumn(format=fmt)

    st.dataframe(disp, hide_index=True, use_container_width=True, column_config=col_cfg)


def tab_overview(df, kpi, groupby2):
    st.subheader("포트폴리오 개요")

    stock_chg = (df["최종포트"] - df["EX_W"])

    def top_mover(sector):
        sub = df[df["GICS"] == sector]
        sub_chg = stock_chg.loc[sub.index]
        return sub.loc[sub_chg.abs().idxmax(), "Name"], sub_chg.loc[sub_chg.abs().idxmax()]

    sector_chg = df.groupby("GICS")[["최종포트", "EX_W"]].sum()
    sector_chg["변화"] = sector_chg["최종포트"] - sector_chg["EX_W"]
    inc_name, inc_chg = sector_chg["변화"].idxmax(), sector_chg["변화"].max()
    dec_name, dec_chg = sector_chg["변화"].idxmin(), sector_chg["변화"].min()
    rest_max = sector_chg["변화"].drop([inc_name, dec_name]).abs().max()
    inc_stock, inc_stock_chg = top_mover(inc_name)
    dec_stock, dec_stock_chg = top_mover(dec_name)
    st.markdown(
        f"- 최종포트가 EX 비중보다 가장 늘어난 업종: **{inc_name}** (+{inc_chg * 100:.1f}%p), "
        f"내 최대 변화 종목: {inc_stock} ({inc_stock_chg * 100:+.1f}%p)\n"
        f"- 최종포트가 EX 비중보다 가장 줄어든 업종: **{dec_name}** ({dec_chg * 100:.1f}%p), "
        f"내 최대 변화 종목: {dec_stock} ({dec_stock_chg * 100:+.1f}%p)\n"
        f"- 나머지 업종은 모두 ±{rest_max * 100:.1f}%p 이내로 변화가 작음"
    )

    aw_chg = df["최종AW"] - df["EX_AW"]
    moved_sectors = sector_chg[sector_chg["변화"].abs() >= 0.02].index.tolist()
    for sec in moved_sectors:
        sec_pct = sector_chg.loc[sec, "변화"] * 100
        sub = df[df["GICS"] == sec]
        sub_aw_chg = aw_chg.loc[sub.index]
        hits = sub.loc[sub_aw_chg[sub_aw_chg.abs() >= 0.02].index]
        if not hits.empty:
            names = ", ".join(
                f"{n} ({c * 100:+.1f}%p)" for n, c in zip(hits["Name"], sub_aw_chg.loc[hits.index])
            )
            st.markdown(f"- **{sec}** (비중 {sec_pct:+.1f}%p) 내 AW가 2%p 이상 변한 종목: {names}")
        else:
            closest_idx = sub_aw_chg.abs().idxmax()
            st.markdown(
                f"- **{sec}** (비중 {sec_pct:+.1f}%p) 내에는 AW가 2%p 이상 변한 종목 없음 "
                f"(최대 변화: {sub.loc[closest_idx, 'Name']} {sub_aw_chg.loc[closest_idx] * 100:+.1f}%p)"
            )

    st.divider()

    kpi_rows = [
        {"항목": "리밸 종목수", "값": f"{int(kpi['리밸_종목수']['value'])}"},
        {"항목": "현재 종목수", "값": f"{int(kpi['현재_종목수']['value'])}"},
        {"항목": f"Inter 복제율 (기준 {fmt_pct(0.50)})", "값": fmt_pct(kpi["Inter_복제율"]["value"])},
        {"항목": "ExPort 복제율", "값": fmt_pct(kpi["ExPort_복제율"]["value"])},
        {"항목": "턴오버", "값": fmt_pct(kpi["턴오버"]["value"])},
        {"항목": f"인터브랜드 비중 (기준 {fmt_pct(0.60)})", "값": fmt_pct(kpi["인터브랜드_비중"]["value"])},
        {"항목": f"BM 복제율 (기준 {fmt_pct(0.30)})", "값": fmt_pct(kpi["BM_복제율"]["value"])},
    ]
    kpi_col_cfg = {c: st.column_config.Column(width="medium") for c in ["항목", "값"]}
    st.dataframe(pd.DataFrame(kpi_rows), hide_index=True, use_container_width=True, column_config=kpi_col_cfg)

    st.divider()

    port_df = df[df["편입"]].copy()

    st.markdown("**편입 종목 최종 스코어 순위**")
    rank_df = (
        port_df[["Name", "GICS", "IndustryGroup", "Country",
                 "Total_B", "B", "Total_C", "C", "Total_Q", "Q", "Total_M", "M",
                 "Final_S", "최종포트", "최종AW", "EX_W", "BM_W"]]
        .sort_values("Final_S", ascending=False)
        .reset_index(drop=True)
    )
    rank_df.index += 1
    rank_df.columns = ["종목명", "섹터", "산업군", "국가",
                        "브랜드(Total)", "브랜드", "경쟁력(Total)", "경쟁력",
                        "Quality(Total)", "Quality", "Macro(Total)", "Macro",
                        "최종", "최종포트", "비중(AW)", "EX 비중", "BM 비중"]
    for c in ["최종포트", "비중(AW)", "EX 비중", "BM 비중"]:
        rank_df[c] = rank_df[c].apply(fmt_pct)

    score_cols = ["브랜드(Total)", "브랜드", "경쟁력(Total)", "경쟁력",
                  "Quality(Total)", "Quality", "Macro(Total)", "Macro", "최종"]
    styled = rank_df.style.map(
        score_bar_color, subset=["브랜드", "경쟁력", "Quality", "Macro", "최종"]
    ).format({c: "{:.2f}" for c in score_cols}, na_rep="—")
    rank_col_cfg = {c: st.column_config.Column(width="small") for c in rank_df.columns}
    st.dataframe(styled, height=700, use_container_width=True, column_config=rank_col_cfg)

    st.divider()
    bar3 = ["차이", "Active", "XActive"]
    render_gb_table("R_TR.GICSSector", groupby2["sector"], ["XPORT", "PORT", "차이"], bar3)
    render_gb_table("테마별 비중", groupby2["theme"], ["XPORT", "PORT", "차이"], bar3)
    render_gb_table("R_TR.GICSIndustryGroup", groupby2["industry_group"], ["XPORT", "PORT", "차이"], bar3)
    render_gb_table("R_TR.CoRPriJaryCountry", groupby2["country"], [], ["Active", "XActive"])


def tab_factor(df):
    st.subheader("팩터 스코어 분석")

    port_df = df[df["편입"]].copy()

    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("**섹터별 평균 팩터 스코어 (편입 종목)**")
        sector_score = (
            port_df.groupby("GICS")[["B", "C", "Q", "M", "Final_S"]]
            .mean()
            .reindex([s for s in SECTOR_ORDER if s in port_df["GICS"].unique()])
        )
        fig = px.imshow(
            sector_score.T,
            text_auto=".2f",
            color_continuous_scale="RdYlGn",
            zmin=0, zmax=1,
            aspect="auto",
            labels=dict(x="섹터", y="팩터", color="스코어"),
        )
        fig.update_layout(margin=dict(l=0, r=0, t=10, b=10), height=320)
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("**팩터별 스코어 분포 (편입 종목)**")
        melt = port_df[["Name", "B", "C", "Q", "M"]].melt(
            id_vars="Name", var_name="팩터", value_name="스코어"
        )
        fig2 = px.box(
            melt, x="팩터", y="스코어",
            color="팩터",
            color_discrete_map=FACTOR_COLORS,
            points="all",
        )
        fig2.update_layout(margin=dict(l=0, r=0, t=10, b=10), height=300,
                           showlegend=False, yaxis=dict(range=[-0.05, 1.1]))
        st.plotly_chart(fig2, use_container_width=True)

    with col_r:
        st.markdown("**편입 종목 팩터 구성 (평균)**")
        avg = port_df[["B", "C", "Q", "M"]].mean()
        fig3 = go.Figure(go.Bar(
            x=avg.index.tolist(),
            y=avg.values.tolist(),
            marker_color=[FACTOR_COLORS[k] for k in avg.index],
            text=[f"{v:.3f}" for v in avg.values],
            textposition="outside",
        ))
        fig3.update_layout(
            yaxis=dict(range=[0, 1.2], title="평균 스코어"),
            margin=dict(l=0, r=0, t=10, b=10), height=240,
        )
        st.plotly_chart(fig3, use_container_width=True)

        st.markdown("**스코어별 편입 종목 Top 15**")
        factor_sel = st.selectbox("정렬 기준 팩터", ["Final_S", "B", "C", "Q", "M"], key="factor_sel")
        top_df = (
            port_df.nlargest(15, factor_sel)[["Name", "B", "C", "Q", "M", "Final_S"]]
            .reset_index(drop=True)
        )
        top_df.index += 1

        fig4 = go.Figure()
        for factor, color in FACTOR_COLORS.items():
            if factor == "최종_S":
                continue
            fig4.add_trace(go.Bar(
                name=factor,
                x=top_df["Name"],
                y=top_df[factor],
                marker_color=color,
            ))
        fig4.update_layout(
            barmode="group",
            margin=dict(l=0, r=0, t=10, b=80),
            height=380,
            xaxis=dict(tickangle=-40, tickfont=dict(size=10)),
            yaxis=dict(range=[0, 1.2]),
            legend=dict(orientation="h", y=1.05),
        )
        st.plotly_chart(fig4, use_container_width=True)


def tab_sector_country(df, groupby2):
    st.subheader("섹터 / 국가 구성")

    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("**섹터별 포트 비중 vs URTH(벤치마크) 비중**")
        sec = (
            groupby2["sector"]
            .set_index("항목")
            .reindex([s for s in SECTOR_ORDER if s in groupby2["sector"]["항목"].unique()])
            .fillna(0)
            .reset_index()
            .rename(columns={"항목": "GICS"})
        )

        fig = go.Figure()
        fig.add_trace(go.Bar(name="포트 비중", x=sec["GICS"], y=sec["PORT비중"],
                             marker_color="#4C72B0"))
        fig.add_trace(go.Bar(name="URTH 비중", x=sec["GICS"], y=sec["URTH"],
                             marker_color="#aec7e8"))
        fig.update_layout(
            barmode="group",
            yaxis=dict(tickformat=".1%"),
            margin=dict(l=0, r=0, t=10, b=100),
            height=380,
            xaxis=dict(tickangle=-40),
            legend=dict(orientation="h", y=1.05),
        )
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("**Over/Underweight (Active = 포트 - URTH)**")
        sec_sorted = sec.sort_values("Active", ascending=True)
        colors = ["#d62728" if v < 0 else "#2ca02c" for v in sec_sorted["Active"]]
        fig2 = go.Figure(go.Bar(
            x=sec_sorted["Active"],
            y=sec_sorted["GICS"],
            orientation="h",
            marker_color=colors,
            text=[f"{v*100:+.2f}%" for v in sec_sorted["Active"]],
            textposition="outside",
        ))
        fig2.update_layout(
            margin=dict(l=0, r=80, t=10, b=10), height=320,
            xaxis=dict(tickformat=".1%"),
        )
        st.plotly_chart(fig2, use_container_width=True)

    with col_r:
        st.markdown("**국가별 포트 비중**")
        country = groupby2["country"].sort_values("PORT비중", ascending=False).reset_index(drop=True)
        if len(country) > 10:
            country_pie = pd.concat([
                country.head(10)[["항목", "PORT비중"]],
                pd.DataFrame([{"항목": "기타", "PORT비중": country.iloc[10:]["PORT비중"].sum()}]),
            ], ignore_index=True)
        else:
            country_pie = country[["항목", "PORT비중"]]

        fig3 = px.pie(country_pie, names="항목", values="PORT비중",
                      hole=0.4,
                      color_discrete_sequence=px.colors.qualitative.Set2)
        fig3.update_traces(textposition="inside", textinfo="percent+label")
        fig3.update_layout(margin=dict(l=0, r=0, t=10, b=10), height=320,
                           showlegend=True,
                           legend=dict(orientation="v", x=1.02))
        st.plotly_chart(fig3, use_container_width=True)

        render_gb_table("국가별 상세 비중", country, [], ["Active", "XActive"])


def tab_themes(df, themes):
    st.subheader("테마 집합 분석 (AGG & ECALL)")

    # 테마별 편입 현황 요약
    summary_rows = []
    for theme_name, tdata in themes.items():
        stocks_df = tdata["stocks"]
        if stocks_df.empty:
            continue
        ecall_stocks = stocks_df[stocks_df["ECALL"] == 1]["ISIN"].dropna().tolist()
        in_port = df[df["편입"] & df["ISIN"].isin(ecall_stocks)].shape[0] if ecall_stocks else 0
        summary_rows.append({
            "테마": theme_name,
            "유니버스 종목수": len(stocks_df),
            "ECALL 종목수": int(tdata["agg_ecall"]) if isinstance(tdata["agg_ecall"], (int, float)) else 0,
            "포트 편입수": in_port,
        })

    if summary_rows:
        sum_df = pd.DataFrame(summary_rows)
        fig = go.Figure(data=[
            go.Bar(name="유니버스", x=sum_df["테마"], y=sum_df["유니버스 종목수"],
                   marker_color="#aec7e8"),
            go.Bar(name="ECALL", x=sum_df["테마"], y=sum_df["ECALL 종목수"],
                   marker_color="#ffbb78"),
            go.Bar(name="포트 편입", x=sum_df["테마"], y=sum_df["포트 편입수"],
                   marker_color="#4C72B0"),
        ])
        fig.update_layout(
            barmode="group",
            margin=dict(l=0, r=0, t=10, b=80),
            height=320,
            xaxis=dict(tickangle=-20),
            legend=dict(orientation="h", y=1.05),
        )
        st.plotly_chart(fig, use_container_width=True)

    st.divider()

    # 테마별 종목 테이블
    theme_names = list(themes.keys())
    if not theme_names:
        st.info("테마 데이터 없음")
        return

    sel_theme = st.selectbox("테마 선택", theme_names, key="theme_sel")
    tdata = themes[sel_theme]
    stocks_df = tdata["stocks"].copy()

    col1, col2 = st.columns(2)
    with col1:
        st.metric("총 종목수", len(stocks_df))
    with col2:
        ecall_cnt = int(tdata["agg_ecall"]) if isinstance(tdata["agg_ecall"], (int, float)) else 0
        st.metric("ECALL 편입 (테마 기준)", ecall_cnt)

    # 포트 편입 여부 조인
    port_isins = set(df[df["편입"]]["ISIN"].tolist())
    stocks_df["포트편입"] = stocks_df["ISIN"].apply(
        lambda x: "✅" if isinstance(x, str) and x in port_isins else "—"
    )

    # 스코어 조인
    score_map = df.set_index("ISIN")[["B", "C", "Q", "M", "Final_S"]].to_dict("index")
    for col in ["B", "C", "Q", "M", "Final_S"]:
        stocks_df[col] = stocks_df["ISIN"].apply(
            lambda x: score_map.get(x, {}).get(col) if isinstance(x, str) else None
        )

    disp = stocks_df[["Name", "AGG", "ECALL", "포트편입", "B", "C", "Q", "M", "Final_S"]].copy()
    disp.columns = ["종목명", "AGG", "ECALL", "포트편입", "브랜드", "경쟁력", "Quality", "Macro", "최종"]
    disp = disp.reset_index(drop=True)
    disp.index += 1

    styled = disp.style.map(
        score_bar_color, subset=["브랜드", "경쟁력", "Quality", "Macro", "최종"]
    ).format(
        {"브랜드": "{:.2f}", "경쟁력": "{:.2f}", "Quality": "{:.2f}", "Macro": "{:.2f}", "최종": "{:.2f}"},
        na_rep="—"
    )
    st.dataframe(styled, height=480, use_container_width=True)


def tab_detail(df):
    st.subheader("종목 상세 검색")

    # 필터
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        sector_opt = ["전체"] + sorted([x for x in df["GICS"].dropna().unique() if isinstance(x, str)])
        sel_sector = st.selectbox("섹터", sector_opt, key="d_sector")
    with col2:
        country_opt = ["전체"] + sorted([x for x in df["Country"].dropna().unique() if isinstance(x, str)])
        sel_country = st.selectbox("국가", country_opt, key="d_country")
    with col3:
        port_opt = {"전체": None, "편입 종목만": True, "미편입만": False}
        sel_port = st.selectbox("편입 여부", list(port_opt.keys()), key="d_port")
    with col4:
        min_score = st.slider("최종 스코어 최소값", 0.0, 1.0, 0.0, 0.05, key="d_score")

    filtered = df.copy()
    if sel_sector != "전체":
        filtered = filtered[filtered["GICS"] == sel_sector]
    if sel_country != "전체":
        filtered = filtered[filtered["Country"] == sel_country]
    if port_opt[sel_port] is not None:
        filtered = filtered[filtered["편입"] == port_opt[sel_port]]
    filtered = filtered[filtered["Final_S"].fillna(0) >= min_score]

    st.markdown(f"**검색 결과: {len(filtered)}개 종목**")

    col_l, col_r = st.columns([1.5, 1])

    with col_l:
        show_df = (
            filtered[["Name", "GICS", "Country", "편입", "B", "C", "Q", "M", "Final_S", "최종AW", "BM_W"]]
            .sort_values("Final_S", ascending=False)
            .reset_index(drop=True)
        )
        show_df.index += 1
        show_df["편입"] = show_df["편입"].map({True: "✅", False: "—"})
        show_df["최종AW"] = show_df["최종AW"].apply(fmt_pct)
        show_df["BM_W"] = show_df["BM_W"].apply(fmt_pct)
        show_df.columns = ["종목명", "섹터", "국가", "편입", "B", "C", "Q", "M", "최종", "AW", "BM W"]

        styled = show_df.style.map(
            score_bar_color, subset=["B", "C", "Q", "M", "최종"]
        ).format({"B": "{:.2f}", "C": "{:.2f}", "Q": "{:.2f}", "M": "{:.2f}", "최종": "{:.2f}"}, na_rep="—")
        st.dataframe(styled, height=560, use_container_width=True)

    with col_r:
        # 종목 선택해서 레이더 차트
        st.markdown("**종목 선택 → 팩터 레이더**")
        name_list = filtered.sort_values("Final_S", ascending=False)["Name"].dropna().tolist()
        if name_list:
            sel_name = st.selectbox("종목 선택", name_list, key="d_name")
            row = filtered[filtered["Name"] == sel_name].iloc[0]

            factors = ["B", "C", "Q", "M"]
            values = [row[f] if pd.notna(row[f]) else 0 for f in factors]
            values_closed = values + [values[0]]
            factors_closed = factors + [factors[0]]

            fig = go.Figure(go.Scatterpolar(
                r=values_closed,
                theta=factors_closed,
                fill="toself",
                fillcolor="rgba(76, 114, 176, 0.3)",
                line=dict(color="#4C72B0", width=2),
                name=sel_name,
            ))
            fig.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 1])),
                margin=dict(l=40, r=40, t=60, b=40),
                height=340,
                title=dict(text=sel_name, font=dict(size=13)),
            )
            st.plotly_chart(fig, use_container_width=True)

            # 세부 정보
            info = {
                "섹터": row["GICS"],
                "산업군": row["IndustryGroup"],
                "국가": row["Country"],
                "편입 여부": "✅ 편입" if row["편입"] else "— 미편입",
                "브랜드(B)": f"{row['B']:.3f}" if pd.notna(row["B"]) else "—",
                "경쟁력(C)": f"{row['C']:.3f}" if pd.notna(row["C"]) else "—",
                "Quality(Q)": f"{row['Q']:.3f}" if pd.notna(row["Q"]) else "—",
                "Macro(M)": f"{row['M']:.3f}" if pd.notna(row["M"]) else "—",
                "최종 스코어": f"{row['Final_S']:.3f}" if pd.notna(row["Final_S"]) else "—",
                "최종 AW": fmt_pct(row["최종AW"]),
                "BM 비중": fmt_pct(row["BM_W"]),
            }
            for k, v in info.items():
                st.markdown(f"**{k}**: {v}")


# ─── 메인 ────────────────────────────────────────────────────────────────────

def tab_portfolio_returns(df, factor_detail):
    st.subheader("편입 종목 비중 & 수익률")

    port_df = df[df["편입"]].copy()

    extra_holdings = pd.DataFrame([
        {"ISIN": "US81369Y1001", "Name": "Materials Select Sector SPDR Fund (XLB)",
         "GICS": "Materials", "Country": "United States", "최종포트": 0.01, "최종AW": 0.01},
        {"ISIN": "US81369Y5069", "Name": "Energy Select Sector SPDR Fund (XLE)",
         "GICS": "Energy", "Country": "United States", "최종포트": 0.01, "최종AW": 0.01},
    ])
    port_df = pd.concat([port_df, extra_holdings], ignore_index=True)

    ticker_map = dict(factor_detail.get("ticker_map", {}))
    ticker_map["US81369Y1001"] = "XLB"
    ticker_map["US81369Y5069"] = "XLE"

    # 편입 종목 중 ticker 있는 것만 fetch
    port_isins = set(port_df["ISIN"].tolist())
    ticker_tuples = tuple(
        sorted((isin, t) for isin, t in ticker_map.items() if isin in port_isins)
    )

    # ── 갱신 버튼 + 데이터 로드 ──
    col_btn, col_info = st.columns([1, 5])
    with col_btn:
        if st.button("🔄 실시간 갱신", key="pr_refresh"):
            fetch_live_returns.clear()
            st.rerun()
    with col_info:
        st.caption(f"Yahoo Finance 실시간 (15분 캐시) · 티커 매핑: {len(ticker_tuples)}개 / 편입 {len(port_isins)}개")

    with st.spinner("Yahoo Finance에서 수익률 데이터 로딩 중..."):
        live = fetch_live_returns(ticker_tuples)

    # live 데이터 → DataFrame
    live_rows = []
    for isin in port_isins:
        rec = live.get(isin, {})
        live_rows.append({
            "ISIN": isin,
            "live_price": rec.get("last_price"),
            "D_R":   rec.get("D"),
            "1W_R":  rec.get("1W"),
            "1M_R":  rec.get("1M"),
            "YTD_R": rec.get("YTD"),
            "ticker": rec.get("ticker", ticker_map.get(isin, "—")),
            "live_error": rec.get("error"),
        })
    live_df = pd.DataFrame(live_rows)

    # 편입 종목 + 라이브 수익률 조인
    merged = port_df.merge(live_df, on="ISIN", how="left")

    # 수익률이 없으면 Traditional Q 시트 값으로 fallback
    df_q = factor_detail["Q"]["df"][["ISIN", "현재가", "YTD_R", "1M_R", "1W_R"]].copy()
    df_q = df_q.drop_duplicates(subset=["ISIN"])
    df_q = df_q.rename(columns={"현재가": "static_price", "YTD_R": "s_YTD", "1M_R": "s_1M", "1W_R": "s_1W"})
    merged = merged.merge(df_q, on="ISIN", how="left")

    for col_live, col_static in [("YTD_R", "s_YTD"), ("1M_R", "s_1M"), ("1W_R", "s_1W")]:
        merged[col_live] = merged[col_live].combine_first(merged[col_static])
    merged["live_price"] = merged["live_price"].combine_first(merged["static_price"])

    # ── AW × 수익률 ──
    merged["AWxD"] = merged["최종AW"] * merged["D_R"]
    merged["AWx1W"] = merged["최종AW"] * merged["1W_R"]
    merged["AWx1M"] = merged["최종AW"] * merged["1M_R"]
    merged["AWxYTD"] = merged["최종AW"] * merged["YTD_R"]

    def small_metric(col, label, v, bold=False):
        color = "#2ca02c" if v >= 0 else "#d62728"
        arrow = "▲" if v >= 0 else "▼"
        if bold:
            col.markdown(
                f"<div style='font-size:14px;font-weight:700;color:#888'>{label}</div>"
                f"<div style='font-size:20px;font-weight:900;color:{color};"
                f"background:{color}1a;border-radius:6px;padding:2px 8px;display:inline-block'>"
                f"{arrow} {v*100:+.2f}%</div>",
                unsafe_allow_html=True,
            )
        else:
            col.markdown(
                f"<div style='font-size:14px;color:#888'>{label}</div>"
                f"<div style='font-size:16px;font-weight:600;color:{color}'>{arrow} {v*100:+.2f}%</div>",
                unsafe_allow_html=True,
            )

    periods = ["일간", "1W", "1M", "YTD"]
    port_vals = [merged["AWxD"].sum(), merged["AWx1W"].sum(), merged["AWx1M"].sum(), merged["AWxYTD"].sum()]

    c1, c2, c3, c4 = st.columns(4)
    for col, label, v in zip([c1, c2, c3, c4], periods, port_vals):
        small_metric(col, f"AW×{label} 합계", v)

    # ── BM(URTH) 대비 상대수익률 ──
    bm_live = fetch_live_returns((("US4642863926", "URTH"),))
    bm_rec = bm_live.get("US4642863926", {})
    bm_vals = [bm_rec.get("D") or 0, bm_rec.get("1W") or 0, bm_rec.get("1M") or 0, bm_rec.get("YTD") or 0]
    rel_vals = [p - b for p, b in zip(port_vals, bm_vals)]

    r1, r2, r3, r4 = st.columns(4)
    for col, label, v in zip([r1, r2, r3, r4], periods, rel_vals):
        small_metric(col, f"상대수익률({label})", v, bold=True)
    st.caption("BM: iShares MSCI World ETF (URTH, US4642863926)")

    st.divider()

    sector_opt = ["전체"] + sorted([s for s in merged["GICS"].dropna().unique() if isinstance(s, str)])
    sel_sector = st.selectbox("섹터 필터", sector_opt, key="pr_sector")

    filtered = merged if sel_sector == "전체" else merged[merged["GICS"] == sel_sector]

    disp = filtered[["Name", "GICS", "ticker", "최종포트", "최종AW", "live_price",
                      "AWxD", "D_R", "AWx1W", "1W_R", "AWx1M", "1M_R", "AWxYTD", "YTD_R"]].copy()
    disp = disp.sort_values("최종포트", ascending=False, na_position="last").reset_index(drop=True)
    disp.index += 1
    disp.columns = ["종목명", "섹터", "티커", "포트비중", "AW", "현재가",
                     "AW×일간", "일간", "AW×1W", "1W", "AW×1M", "1M", "AW×YTD", "YTD"]

    # ret_color는 숫자값에 적용 (format 전에), NaN·비숫자 안전 처리
    def ret_color(v):
        if not isinstance(v, (int, float)) or pd.isna(v):
            return ""
        return "color:#2ca02c" if v > 0 else "color:#d62728" if v < 0 else ""

    def _fmt_ret(v):
        return "—" if (not isinstance(v, (int, float)) or pd.isna(v)) else f"{v*100:+.2f}%"

    def _fmt_w(v):
        return "—" if (not isinstance(v, (int, float)) or pd.isna(v)) else f"{v*100:.2f}%"

    def _fmt_price(v):
        return "—" if (not isinstance(v, (int, float)) or pd.isna(v)) else f"{v:,.2f}"

    ret_cols = ["AW×일간", "일간", "AW×1W", "1W", "AW×1M", "1M", "AW×YTD", "YTD"]
    styled = (
        disp.style
        .format({
            "포트비중": _fmt_w,
            "AW": _fmt_w,
            "현재가": _fmt_price,
            **{c: _fmt_ret for c in ret_cols},
        })
        .map(ret_color, subset=ret_cols)
    )
    st.markdown(
        f"<div style='font-size:16px;max-height:580px;overflow-y:auto'>{styled.to_html()}</div>",
        unsafe_allow_html=True,
    )

    st.markdown("**일간 ±5% 이상 급등/급락 종목**")
    movers = merged[merged["D_R"].abs() >= 0.05][["ISIN", "Name", "ticker", "D_R"]].copy()
    movers = movers.sort_values("D_R", key=lambda s: s.abs(), ascending=False)

    if movers.empty:
        st.caption("오늘 ±5% 이상 변동한 종목 없음")
    else:
        for _, row in movers.iterrows():
            color = "#2ca02c" if row["D_R"] > 0 else "#d62728"
            arrow = "▲" if row["D_R"] > 0 else "▼"
            st.markdown(
                f"**{row['Name']}** ({row['ticker']}) "
                f"<span style='color:{color};font-weight:700'>{arrow} {row['D_R']*100:+.2f}%</span>",
                unsafe_allow_html=True,
            )

            moat = top_moat_factor(factor_detail, row["ISIN"])
            profile = fetch_company_profile(row["ticker"]) if isinstance(row["ticker"], str) else {}
            summary = (profile.get("summary") or "").split(". ")[0]
            industry = profile.get("industry")
            overview_bits = [b for b in [industry, summary] if b]
            moat_bit = f"주요 모트 요인: {moat}" if moat else "주요 모트 요인: —"
            if overview_bits:
                st.caption(f"{' · '.join(overview_bits)} · {moat_bit}")
            else:
                st.caption(moat_bit)

            news_items = fetch_stock_news(row["ticker"], count=1) if isinstance(row["ticker"], str) else []
            if news_items:
                item = news_items[0]
                st.markdown(f"- [{item['title']}]({item['link']}) · {item['publisher']}")
            else:
                st.caption("관련 뉴스 없음")

    # 데이터 소스 및 에러 리포트
    errors = [(r.get("ticker", "—"), r.get("error")) for r in live.values() if r.get("error")]
    if errors:
        with st.expander(f"⚠️ 라이브 데이터 조회 실패 종목 ({len(errors)}개)"):
            err_df = pd.DataFrame(errors, columns=["티커", "오류"])
            st.dataframe(err_df, hide_index=True, use_container_width=True)


def tab_factor_detail(df, factor_detail):
    st.subheader("팩터 세부 항목 분석")

    factor_sel = st.radio(
        "팩터 선택", ["B (브랜드)", "C (경쟁력)", "Q (Traditional Q)", "M (Macro)"],
        horizontal=True, key="fd_factor"
    )
    factor_key = factor_sel[0]
    fdata = factor_detail[factor_key]
    detail_df = fdata["df"].copy()

    # 포트 편입 종목만
    port_isins = set(df[df["편입"]]["ISIN"].tolist())
    port_names = df[df["편입"]].set_index("ISIN")["Name"].to_dict()

    detail_df = detail_df[detail_df["ISIN"].isin(port_isins)].copy()
    detail_df = detail_df.drop_duplicates(subset=["ISIN"])
    detail_df["Name"] = detail_df["ISIN"].map(port_names)
    detail_df = detail_df.dropna(subset=["Name"])

    dup_mask = detail_df["Name"].duplicated(keep=False)
    detail_df.loc[dup_mask, "Name"] = (
        detail_df.loc[dup_mask, "Name"] + " (" + detail_df.loc[dup_mask, "ISIN"].str[-4:] + ")"
    )

    items = fdata["items"]
    weights = fdata["weights"]

    col_l, col_r = st.columns([1.6, 1])

    with col_l:
        st.markdown(f"**편입 종목 세부 항목 히트맵** ({fdata['label']})")

        heatmap_df = detail_df.set_index("Name")[items].copy()
        heatmap_df = heatmap_df.fillna(0)

        # 정규화 (0~1 범위로)
        for col in items:
            mx = heatmap_df[col].max()
            if mx > 1:
                heatmap_df[col] = heatmap_df[col] / mx

        # 최종 스코어 기준 정렬
        score_map = df.set_index("ISIN")[["Final_S", "B", "C", "Q", "M"]].to_dict("index")
        detail_df["Final_S"] = detail_df["ISIN"].map(lambda x: score_map.get(x, {}).get("Final_S"))
        heatmap_df = heatmap_df.loc[
            detail_df.set_index("Name")["Final_S"].sort_values(ascending=False).index
        ]

        fig = px.imshow(
            heatmap_df.T,
            text_auto=".2f",
            color_continuous_scale="RdYlGn",
            zmin=0, zmax=1,
            aspect="auto",
            labels=dict(x="종목", y="세부 항목", color="스코어"),
        )
        fig.update_layout(
            margin=dict(l=0, r=0, t=10, b=80),
            height=max(280, len(items) * 55),
            xaxis=dict(tickangle=-45, tickfont=dict(size=9)),
        )
        st.plotly_chart(fig, use_container_width=True)

        # 가중치 표
        st.markdown("**항목별 가중치**")
        w_df = pd.DataFrame([
            {"세부 항목": k, "가중치": f"{v*100:.0f}%"}
            for k, v in weights.items()
        ])
        st.dataframe(w_df, hide_index=True, use_container_width=False)

    with col_r:
        st.markdown("**종목 선택 → 세부 레이더**")
        name_list = detail_df.sort_values("Final_S", ascending=False)["Name"].dropna().tolist()
        sel_name = st.selectbox("종목", name_list, key="fd_name")

        if sel_name:
            row = detail_df[detail_df["Name"] == sel_name].iloc[0]
            vals = []
            for it in items:
                v = row[it]
                mx = detail_df[it].max()
                norm = (v / mx) if (pd.notna(v) and mx > 0) else 0
                vals.append(round(float(norm), 3))

            vals_closed = vals + [vals[0]]
            items_closed = items + [items[0]]

            fig2 = go.Figure(go.Scatterpolar(
                r=vals_closed,
                theta=items_closed,
                fill="toself",
                fillcolor=f"rgba({','.join(str(int(c,16)) for c in ['4C','72','B0'])},0.3)"
                    if factor_key == "B" else
                    f"rgba(221,132,82,0.3)" if factor_key == "C" else
                    f"rgba(85,168,104,0.3)" if factor_key == "Q" else
                    f"rgba(196,78,82,0.3)",
                line=dict(color=FACTOR_COLORS[factor_key], width=2),
                name=sel_name,
            ))
            fig2.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 1])),
                margin=dict(l=40, r=40, t=50, b=20),
                height=340,
                title=dict(text=sel_name[:30], font=dict(size=12)),
            )
            st.plotly_chart(fig2, use_container_width=True)

            # 세부값 테이블
            st.markdown("**세부 항목 원본값**")
            raw_items = items.copy()
            if factor_key == "C":
                raw_items = fdata["moat_items"] + items
            elif factor_key == "M":
                raw_items = fdata["theme_items"] + items

            rows_list = []
            for it in raw_items:
                v = row[it] if it in row.index else None
                w = weights.get(it)
                rows_list.append({
                    "항목": it,
                    "값": f"{v:.3f}" if pd.notna(v) else "—",
                    "가중치": f"{w*100:.0f}%" if w else "—",
                })
            st.dataframe(pd.DataFrame(rows_list), hide_index=True, use_container_width=True)


def check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True

    st.title("🔒 03Y51 스코어보드")
    pw = st.text_input("비밀번호", type="password")
    if st.button("입력"):
        correct = st.secrets.get("password")
        if correct and pw == correct:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")
    return False


def main():
    if not check_password():
        return

    st.title("03Y51 스코어보드 대시보드")
    st.caption("03Y51 운용파일_20260602(6월)_mk.xlsx 기준")

    try:
        df, kpi, themes, factor_detail, groupby2 = load_data()
    except Exception as e:
        st.error(f"데이터 로드 실패: {e}")
        return

    tabs = st.tabs([
        "💹 비중 & 수익률",
        "📋 포트폴리오 개요",
        "📊 팩터 스코어 분석",
        "🗺️ 섹터/국가 구성",
        "🎯 테마 집합 분석",
        "🔬 팩터 세부 분석",
        "🔍 종목 상세 검색",
    ])

    import traceback
    with tabs[0]:
        try:
            tab_portfolio_returns(df, factor_detail)
        except Exception as e:
            st.error(f"[탭1 에러] {e}")
            st.code(traceback.format_exc())
    with tabs[1]:
        try:
            tab_overview(df, kpi, groupby2)
        except Exception as e:
            st.error(f"[탭2 에러] {e}")
            st.code(traceback.format_exc())
    with tabs[2]:
        try:
            tab_factor(df)
        except Exception as e:
            st.error(f"[탭3 에러] {e}")
            st.code(traceback.format_exc())
    with tabs[3]:
        try:
            tab_sector_country(df, groupby2)
        except Exception as e:
            st.error(f"[탭4 에러] {e}")
            st.code(traceback.format_exc())
    with tabs[4]:
        try:
            tab_themes(df, themes)
        except Exception as e:
            st.error(f"[탭5 에러] {e}")
            st.code(traceback.format_exc())
    with tabs[5]:
        try:
            tab_factor_detail(df, factor_detail)
        except Exception as e:
            st.error(f"[탭6 에러] {e}")
            st.code(traceback.format_exc())
    with tabs[6]:
        try:
            tab_detail(df)
        except Exception as e:
            st.error(f"[탭7 에러] {e}")
            st.code(traceback.format_exc())


if __name__ == "__main__":
    main()
